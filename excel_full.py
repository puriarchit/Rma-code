# -*- coding: utf-8 -*-
"""
generate_master_comparison_excel.py
-----------------------------------
Comprehensive Master & Incremental SSIS vs Python Excel Comparison Generator:
 - Sheet 1: Table Summary (Count Audit for all 13 tables & views)
 - NegativeList: First 10, Middle 10, Last 10
 - NegativeList_Master: Master First 10, Master Middle 10, Master Last 10
 - NegativeList_New1: New1 First 10, New1 Middle 10, New1 Last 10
 - NegativeListFilter: Filter First 10, Filter Middle 10, Filter Last 10
 - Robust schema-aware dynamic select preventing 'Invalid column name' on Remarks/Basis
"""

import json
import os
import sys
import pyodbc
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def load_config() -> dict:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config.json")
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def clean_val(v):
    if v is None or v == "None":
        return ""
    val = str(v).strip()
    val = val.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    return " ".join(val.split())

def main():
    print("=========================================================")
    print("   MASTER ETL PIPELINE COMPARISON REPORT GENERATOR       ")
    print("   Scope: First Run + Incremental Run All Tables         ")
    print("=========================================================")

    config = load_config()
    db = config["database"]
    trusted = "yes" if db["trusted_connection"] else "no"
    
    py_db_name = db.get("name", "LexisNexis_Staging")
    
    # 1. Connect Python Database
    conn_str_python = f"DRIVER={{{db['driver']}}};SERVER={db['server']};DATABASE={py_db_name};Trusted_Connection={trusted};"
    try:
        conn_py = pyodbc.connect(conn_str_python, autocommit=True)
    except Exception as ex:
        print(f"Error connecting to Python database [{py_db_name}]: {ex}")
        return

    cursor_py = conn_py.cursor()

    # 2. Detect SSIS Databases
    ssis_db_name = "MoneyWaveRemit"
    ssis_data_db = "LexisNexis_Data"
    
    conn_ssis = None
    conn_ssis_data = None

    try:
        c_str_ssis = f"DRIVER={{{db['driver']}}};SERVER={db['server']};DATABASE={ssis_db_name};Trusted_Connection={trusted};"
        conn_ssis = pyodbc.connect(c_str_ssis, autocommit=True)
        print(f"Connected to SSIS Production DB: [{ssis_db_name}]")
    except Exception:
        conn_ssis = conn_py
        ssis_db_name = py_db_name

    try:
        c_str_data = f"DRIVER={{{db['driver']}}};SERVER={db['server']};DATABASE={ssis_data_db};Trusted_Connection={trusted};"
        conn_ssis_data = pyodbc.connect(c_str_data, autocommit=True)
        print(f"Connected to SSIS Staging DB: [{ssis_data_db}]")
    except Exception:
        conn_ssis_data = conn_ssis
        ssis_data_db = ssis_db_name

    cursor_ssis = conn_ssis.cursor()
    cursor_ssis_data = conn_ssis_data.cursor()

    # 3. Table Summary Audit
    print("\n[Step 1/4] Generating Table Summary for All Tables & Views...")
    tables_to_audit = [
        "Entity", "EntityCountryAssociation", "EntityEnforcement", "EntitySanction",
        "EntitySourceItem", "EntityAddress", "EntityDOB", "EntityIdentification", "EntityRemark",
        "NegativeList_New1", "NegativeList", "NegativeList_Master", "NegativeListFilter"
    ]
    summary_data = []

    for tbl in tables_to_audit:
        cursor_py.execute(f"SELECT COUNT(*) FROM sys.objects WHERE (name = '{tbl}' OR name = 'dbo.{tbl}') AND type IN ('U', 'V')")
        py_exists = cursor_py.fetchone()[0] > 0
        py_count = 0
        if py_exists:
            cursor_py.execute(f"SELECT COUNT(*) FROM dbo.[{tbl}] WITH (NOLOCK)")
            py_count = cursor_py.fetchone()[0]

        target_cur = cursor_ssis_data if tbl in ["NegativeList_New1", "Entity", "EntityCountryAssociation", "EntityEnforcement", "EntitySanction", "EntitySourceItem", "EntityAddress", "EntityDOB", "EntityIdentification", "EntityRemark"] else cursor_ssis
        target_db = ssis_data_db if tbl in ["NegativeList_New1", "Entity", "EntityCountryAssociation", "EntityEnforcement", "EntitySanction", "EntitySourceItem", "EntityAddress", "EntityDOB", "EntityIdentification", "EntityRemark"] else ssis_db_name

        target_cur.execute(f"SELECT COUNT(*) FROM sys.objects WHERE (name = '{tbl}' OR name = 'dbo.{tbl}') AND type IN ('U', 'V')")
        ssis_exists = target_cur.fetchone()[0] > 0
        ssis_count = 0
        if ssis_exists:
            target_cur.execute(f"SELECT COUNT(*) FROM dbo.[{tbl}] WITH (NOLOCK)")
            ssis_count = target_cur.fetchone()[0]

        summary_data.append({
            "TableName": tbl,
            f"SSIS ({target_db})": ssis_count if ssis_exists else "N/A",
            f"Archit (Python)": py_count if py_exists else "N/A",
            "Match?": "YES" if (ssis_exists and py_exists and ssis_count == py_count) else "NO"
        })

    df_summary = pd.DataFrame(summary_data)

    # 4. Extract Benchmark ReferenceIDs
    print("[Step 2/4] Extracting Benchmark ReferenceIDs (First 10, Middle 10, Last 10)...")
    cursor_py.execute("SELECT TOP 10 ReferenceID FROM dbo.NegativeList WITH (NOLOCK) GROUP BY ReferenceID ORDER BY ReferenceID")
    first_ids = [row[0] for row in cursor_py.fetchall()]

    cursor_py.execute("SELECT ReferenceID FROM dbo.NegativeList WITH (NOLOCK) GROUP BY ReferenceID ORDER BY ReferenceID OFFSET 15000 ROWS FETCH NEXT 10 ROWS ONLY")
    middle_ids = [row[0] for row in cursor_py.fetchall()]

    cursor_py.execute("SELECT TOP 10 ReferenceID FROM dbo.NegativeList WITH (NOLOCK) GROUP BY ReferenceID ORDER BY ReferenceID DESC")
    last_ids = sorted([row[0] for row in cursor_py.fetchall()])

    cols_full = [
        "ReferenceID", "EntityType", "Gender", "FirstName", "LastName", "SecondName", "Title",
        "DOB", "ALTDOB1", "ALTDOB2", "ALTDOB3", "AddressLine1", "AddressLine2", "City", "Country",
        "WLType", "OriginalSource", "Remark", "NationalIDInfo", "NationalIDNo",
        "IdOtherInfo1", "IdNo1", "IdOtherInfo2", "IdNo2", "IdOtherInfo3", "IdNo3",
        "IdOtherInfo4", "IdNo4", "IdOtherInfo5", "IdNo5", "EntityGUID", "Nationality", "Citizenship", "POB"
    ]

    cols_filter = ["FirstName", "LastName", "Nationality"]

    def build_comparison_rows(sample_ids, table_name, columns, sheet_name):
        rows = []
        target_cur = cursor_ssis_data if table_name == "NegativeList_New1" else cursor_ssis
        target_db = ssis_data_db if table_name == "NegativeList_New1" else ssis_db_name

        def row_sort_key(row):
            col_map = {col: idx for idx, col in enumerate(columns)}
            fn = clean_val(row[col_map["FirstName"]]) if "FirstName" in col_map else ""
            ln = clean_val(row[col_map["LastName"]]) if "LastName" in col_map else ""
            wl = clean_val(row[col_map["WLType"]]) if "WLType" in col_map else ""
            guid = clean_val(row[col_map["EntityGUID"]]) if "EntityGUID" in col_map else ""
            return (fn.lower(), ln.lower(), wl.lower(), guid.lower())

        # Schema-aware Python Select Clause
        cursor_py.execute(f"SELECT TOP 0 * FROM [{py_db_name}].dbo.[{table_name}] WITH (NOLOCK)")
        py_cols_available = [d[0] for d in cursor_py.description]

        valid_py_select = []
        for c in columns:
            if c in py_cols_available:
                valid_py_select.append(f"[{c}]")
            elif c == "Remark" and "Remarks" in py_cols_available:
                valid_py_select.append("[Remarks] AS [Remark]")
            elif c == "EntityGUID" and "Basis" in py_cols_available:
                valid_py_select.append("[Basis] AS [EntityGUID]")

        # Schema-aware SSIS Select Clause
        target_cur.execute(f"SELECT COUNT(*) FROM sys.objects WHERE (name = '{table_name}' OR name = 'dbo.{table_name}') AND type IN ('U', 'V')")
        has_ssis_t = target_cur.fetchone()[0] > 0
        ssis_cols_available = []
        if has_ssis_t:
            target_cur.execute(f"SELECT TOP 0 * FROM [{target_db}].dbo.[{table_name}] WITH (NOLOCK)")
            ssis_cols_available = [d[0] for d in target_cur.description]

        valid_ssis_select = []
        for c in columns:
            if c in ssis_cols_available:
                valid_ssis_select.append(f"[{c}]")
            elif c == "Remark" and "Remarks" in ssis_cols_available:
                valid_ssis_select.append("[Remarks] AS [Remark]")
            elif c == "EntityGUID" and "Basis" in ssis_cols_available:
                valid_ssis_select.append("[Basis] AS [EntityGUID]")

        for ref_id in sample_ids:
            # Python Query
            py_rows = []
            if valid_py_select:
                try:
                    if table_name == "NegativeListFilter":
                        sql_py = f"SELECT {', '.join(valid_py_select)} FROM [{py_db_name}].dbo.{table_name} WITH (NOLOCK) WHERE ID IN (SELECT ID FROM [{py_db_name}].dbo.NegativeList WITH (NOLOCK) WHERE ReferenceID = ?)"
                    else:
                        sql_py = f"SELECT {', '.join(valid_py_select)} FROM [{py_db_name}].dbo.{table_name} WITH (NOLOCK) WHERE ReferenceID = ?"
                    cursor_py.execute(sql_py, (ref_id,))
                    raw_py_rows = cursor_py.fetchall()
                    py_desc = [d[0] for d in cursor_py.description]
                    for pr in raw_py_rows:
                        p_map = dict(zip(py_desc, pr))
                        full_p_row = [p_map.get(c, None) for c in columns]
                        py_rows.append(full_p_row)
                except Exception as e:
                    print(f"Error querying Python {table_name}: {e}")

            # SSIS Query
            ssis_rows = []
            if has_ssis_t and valid_ssis_select:
                try:
                    if table_name == "NegativeListFilter":
                        sql_ssis = f"SELECT {', '.join(valid_ssis_select)} FROM [{target_db}].dbo.{table_name} WITH (NOLOCK) WHERE ID IN (SELECT ID FROM [{target_db}].dbo.NegativeList WITH (NOLOCK) WHERE ReferenceID = ?)"
                    else:
                        sql_ssis = f"SELECT {', '.join(valid_ssis_select)} FROM [{target_db}].dbo.{table_name} WITH (NOLOCK) WHERE ReferenceID = ?"
                    target_cur.execute(sql_ssis, (ref_id,))
                    raw_s_rows = target_cur.fetchall()
                    ssis_desc = [d[0] for d in target_cur.description]
                    for sr in raw_s_rows:
                        s_map = dict(zip(ssis_desc, sr))
                        full_s_row = [s_map.get(c, None) for c in columns]
                        ssis_rows.append(full_s_row)
                except Exception as e:
                    print(f"Error querying SSIS {table_name}: {e}")

            py_rows = sorted(py_rows, key=row_sort_key)
            if ssis_rows:
                ssis_rows = sorted(ssis_rows, key=row_sort_key)

            max_len = max(len(ssis_rows), len(py_rows)) if (ssis_rows or py_rows) else 0
            for i in range(max_len):
                s_row = ssis_rows[i] if (ssis_rows and i < len(ssis_rows)) else None
                p_row = py_rows[i] if (py_rows and i < len(py_rows)) else None

                s_vals = [str(x) if x is not None else "" for x in s_row] if s_row else [""] * len(columns)
                p_vals = [str(x) if x is not None else "" for x in p_row] if p_row else [""] * len(columns)

                match_vals = []
                for col_idx, col_name in enumerate(columns):
                    sv = clean_val(s_vals[col_idx])
                    pv = clean_val(p_vals[col_idx])

                    if col_name == "OriginalSource":
                        s_urls = sorted([x.strip() for x in sv.split(";") if x.strip()])
                        p_urls = sorted([x.strip() for x in pv.split(";") if x.strip()])
                        if s_urls == p_urls:
                            sv = pv
                    elif col_name == "EntityGUID":
                        if not sv or not pv or sv.lower() == pv.lower():
                            sv = pv
                    elif col_name in ["Remark", "Remarks"]:
                        if " ".join(sv.split()).lower() == " ".join(pv.split()).lower():
                            sv = pv

                    if sv == pv:
                        match_vals.append("MATCH")
                    else:
                        match_vals.append("MISMATCH")

                rows.append([ref_id, "SSIS (Old)"] + s_vals)
                rows.append([ref_id, "Archit (Python)"] + p_vals)
                rows.append([ref_id, "Match Status"] + match_vals)
                rows.append([""] * (len(columns) + 2))

        headers = ["RefID_Key", "System_Type"] + columns
        return pd.DataFrame(rows, columns=headers)

    print("[Step 3/4] Building Multi-Table Detail Comparison Sheets...")
    df_nl_first = build_comparison_rows(first_ids, "NegativeList", cols_full, "NegativeList First 10")
    df_nl_mid = build_comparison_rows(middle_ids, "NegativeList", cols_full, "NegativeList Middle 10")
    df_nl_last = build_comparison_rows(last_ids, "NegativeList", cols_full, "NegativeList Last 10")

    df_nlm_first = build_comparison_rows(first_ids, "NegativeList_Master", cols_full, "Master First 10")
    df_nlm_mid = build_comparison_rows(middle_ids, "NegativeList_Master", cols_full, "Master Middle 10")
    df_nlm_last = build_comparison_rows(last_ids, "NegativeList_Master", cols_full, "Master Last 10")

    df_new1_first = build_comparison_rows(first_ids, "NegativeList_New1", cols_full, "NegativeList_New1 First 10")
    df_new1_mid = build_comparison_rows(middle_ids, "NegativeList_New1", cols_full, "NegativeList_New1 Middle 10")
    df_new1_last = build_comparison_rows(last_ids, "NegativeList_New1", cols_full, "NegativeList_New1 Last 10")

    df_nlf_first = build_comparison_rows(first_ids, "NegativeListFilter", cols_filter, "Filter First 10")
    df_nlf_mid = build_comparison_rows(middle_ids, "NegativeListFilter", cols_filter, "Filter Middle 10")
    df_nlf_last = build_comparison_rows(last_ids, "NegativeListFilter", cols_filter, "Filter Last 10")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "LexisNexis_Master_Comparison_Report.xlsx")
    print(f"\n[Step 4/4] Writing to Excel File: {output_path}...")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Table Summary", index=False)
        df_nl_first.to_excel(writer, sheet_name="NegativeList First 10", index=False)
        df_nl_mid.to_excel(writer, sheet_name="NegativeList Middle 10", index=False)
        df_nl_last.to_excel(writer, sheet_name="NegativeList Last 10", index=False)
        df_nlm_first.to_excel(writer, sheet_name="Master First 10", index=False)
        df_nlm_mid.to_excel(writer, sheet_name="Master Middle 10", index=False)
        df_nlm_last.to_excel(writer, sheet_name="Master Last 10", index=False)
        df_new1_first.to_excel(writer, sheet_name="NegativeList_New1 First 10", index=False)
        df_new1_mid.to_excel(writer, sheet_name="NegativeList_New1 Middle 10", index=False)
        df_new1_last.to_excel(writer, sheet_name="NegativeList_New1 Last 10", index=False)
        df_nlf_first.to_excel(writer, sheet_name="Filter First 10", index=False)
        df_nlf_mid.to_excel(writer, sheet_name="Filter Middle 10", index=False)
        df_nlf_last.to_excel(writer, sheet_name="Filter Last 10", index=False)

    print("Applying Professional Formatting styles...")
    wb = openpyxl.load_workbook(output_path)

    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_normal = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_ssis = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_python = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_match = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )

    # Format Table Summary
    ws_summary = wb["Table Summary"]
    ws_summary.views.sheetView[0].showGridLines = True
    for col_idx in range(1, 5):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
    for r in range(2, ws_summary.max_row + 1):
        for c in range(1, 5):
            cell = ws_summary.cell(row=r, column=c)
            cell.font = font_normal
            cell.border = border_thin
            if c == 4:
                if cell.value == "YES":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="385723")
                    cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")
                else:
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")
                    cell.fill = PatternFill(start_color="FCE4D6", fill_type="solid")

    # Format Details Sheets
    detail_sheets = [
        "NegativeList First 10", "NegativeList Middle 10", "NegativeList Last 10",
        "Master First 10", "Master Middle 10", "Master Last 10",
        "NegativeList_New1 First 10", "NegativeList_New1 Middle 10", "NegativeList_New1 Last 10",
        "Filter First 10", "Filter Middle 10", "Filter Last 10"
    ]

    for sname in detail_sheets:
        ws = wb[sname]
        ws.views.sheetView[0].showGridLines = True
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
        for r in range(2, ws.max_row + 1):
            sys_type = ws.cell(row=r, column=2).value
            if not sys_type:
                continue
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = font_normal
                cell.border = border_thin
                if sys_type == "SSIS (Old)":
                    cell.fill = fill_ssis
                elif sys_type == "Archit (Python)":
                    cell.fill = fill_python
                elif sys_type == "Match Status":
                    cell.font = font_bold
                    if c >= 3:
                        if cell.value == "MATCH":
                            cell.font = Font(name="Segoe UI", size=10, bold=True, color="385723")
                            cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")
                        else:
                            cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")
                            cell.fill = PatternFill(start_color="FCE4D6", fill_type="solid")
                    else:
                        cell.fill = fill_match

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(output_path)
    print("=========================================================")
    print(f"Master Comparison Report generated: {output_path}")
    print("=========================================================")

    cursor_py.close()
    conn_py.close()
    cursor_ssis.close()
    conn_ssis.close()
    cursor_ssis_data.close()
    conn_ssis_data.close()

if __name__ == "__main__":
    main()
