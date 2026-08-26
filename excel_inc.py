# -*- coding: utf-8 -*-
import json
import os
import sys

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    os.system(f'"{sys.executable}" -m pip install pandas openpyxl')
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import pyodbc

def load_config() -> dict:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config.json")
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def main():
    print("Initializing Database Connections...")
    config = load_config()
    db = config["database"]
    trusted = "yes" if db["trusted_connection"] else "no"
    
    # Python Database: LexisNexis_Staging
    conn_str_python = f"DRIVER={{{db['driver']}}};SERVER={db['server']};DATABASE=LexisNexis_Staging;Trusted_Connection={trusted};"
    # SSIS Database: MoneyWaveRemit (Production)
    conn_str_ssis = f"DRIVER={{{db['driver']}}};SERVER={db['server']};DATABASE=MoneyWaveRemit;Trusted_Connection={trusted};"
    
    try:
        conn_py = pyodbc.connect(conn_str_python, autocommit=True)
        conn_ssis = pyodbc.connect(conn_str_ssis, autocommit=True)
    except Exception as ex:
        print(f"Error connecting to database: {ex}")
        print("Please check your SQL Server name and permissions.")
        return

    cursor_py = conn_py.cursor()
    cursor_ssis = conn_ssis.cursor()

    # 1. Generate Table Summary Sheet
    print("Generating Table Summary...")
    tables = ["NegativeList", "NegativeListFilter", "NegativeList_Master"]
    summary_data = []
    
    for tbl in tables:
        # Check Python Database
        cursor_py.execute(f"SELECT COUNT(*) FROM sys.objects WHERE name = '{tbl}' AND type IN ('U', 'V')")
        py_exists = cursor_py.fetchone()[0] > 0
        py_count = 0
        if py_exists:
            cursor_py.execute(f"SELECT COUNT(*) FROM dbo.[{tbl}]")
            py_count = cursor_py.fetchone()[0]

        # Check SSIS Database
        cursor_ssis.execute(f"SELECT COUNT(*) FROM sys.objects WHERE name = '{tbl}' AND type IN ('U', 'V')")
        ssis_exists = cursor_ssis.fetchone()[0] > 0
        ssis_count = 0
        if ssis_exists:
            cursor_ssis.execute(f"SELECT COUNT(*) FROM dbo.[{tbl}]")
            ssis_count = cursor_ssis.fetchone()[0]

        summary_data.append({
            "TableName": tbl,
            "SSIS (MoneyWaveRemit)": ssis_count if ssis_exists else "N/A",
            "Python (LexisNexis_Staging)": py_count if py_exists else "N/A",
            "Match?": "YES" if ssis_count == py_count else "NO"
        })
    df_summary = pd.DataFrame(summary_data)

    # Sample ID Extraction from NegativeList to align row-level comparison
    print("Extracting Sample IDs for Row-Level Comparison...")
    cursor_py.execute("SELECT TOP 10 ReferenceID FROM dbo.NegativeList GROUP BY ReferenceID ORDER BY ReferenceID")
    first_ids = [row[0] for row in cursor_py.fetchall()]
    
    cursor_py.execute("SELECT ReferenceID FROM dbo.NegativeList GROUP BY ReferenceID ORDER BY ReferenceID OFFSET 15000 ROWS FETCH NEXT 10 ROWS ONLY")
    middle_ids = [row[0] for row in cursor_py.fetchall()]
    
    cursor_py.execute("SELECT TOP 10 ReferenceID FROM dbo.NegativeList GROUP BY ReferenceID ORDER BY ReferenceID DESC")
    last_ids = sorted([row[0] for row in cursor_py.fetchall()])

    def clean_val(v):
        if v is None or v == "None":
            return ""
        val = str(v).strip()
        # Normalise line endings and spacing to prevent fake mismatches
        val = val.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
        return " ".join(val.split())

    mismatches_summary = []

    # Columns configuration for NegativeList / NegativeList_Master
    cols_negativelist = [
        "ReferenceID", "EntityType", "Gender", "FirstName", "LastName", "SecondName", "Title",
        "DOB", "ALTDOB1", "ALTDOB2", "ALTDOB3", "AddressLine1", "AddressLine2", "City", "Country",
        "WLType", "OriginalSource", "Remark", "NationalIDInfo", "NationalIDNo",
        "IdOtherInfo1", "IdNo1", "IdOtherInfo2", "IdNo2", "IdOtherInfo3", "IdNo3",
        "IdOtherInfo4", "IdNo4", "IdOtherInfo5", "IdNo5", "EntityGUID", "Nationality", "Citizenship", "POB"
    ]

    # Columns configuration for NegativeListFilter
    cols_filter = ["ID", "FirstName", "LastName", "Nationality"]

    # Columns configuration for NegativeList_Master View (uses Remarks and Basis)
    cols_master = [
        "ReferenceID", "EntityType", "Gender", "FirstName", "LastName", "SecondName", "Title",
        "DOB", "ALTDOB1", "ALTDOB2", "ALTDOB3", "AddressLine1", "AddressLine2", "City", "Country",
        "WLType", "OriginalSource", "Remarks", "NationalIDInfo", "NationalIDNo",
        "IdOtherInfo1", "IdNo1", "IdOtherInfo2", "IdNo2", "IdOtherInfo3", "IdNo3",
        "IdOtherInfo4", "IdNo4", "IdOtherInfo5", "IdNo5", "Basis", "Nationality", "Citizenship", "POB"
    ]

    def build_comparison_rows(sample_ids, table_name, columns, sheet_name):
        rows = []
        def row_sort_key(row):
            col_map = {col: idx for idx, col in enumerate(columns)}
            fn = clean_val(row[col_map["FirstName"]]) if "FirstName" in col_map else ""
            ln = clean_val(row[col_map["LastName"]]) if "LastName" in col_map else ""
            wl = clean_val(row[col_map["WLType"]]) if "WLType" in col_map else ""
            guid = ""
            if "EntityAliasGUID" in col_map:
                guid = clean_val(row[col_map["EntityAliasGUID"]])
            elif "Basis" in col_map:
                guid = clean_val(row[col_map["Basis"]])
            elif "ID" in col_map:
                guid = str(row[col_map["ID"]])
            return (fn.lower(), ln.lower(), wl.lower(), guid.lower())

        for ref_id in sample_ids:
            # Query from Python (LexisNexis_Staging)
            if table_name == "NegativeListFilter":
                sql_py = f"SELECT {', '.join(columns)} FROM LexisNexis_Staging.dbo.{table_name} WHERE ID IN (SELECT ID FROM LexisNexis_Staging.dbo.NegativeList WHERE ReferenceID = ?)"
                cursor_py.execute(sql_py, (ref_id,))
            else:
                sql_py = f"SELECT {', '.join(columns)} FROM LexisNexis_Staging.dbo.{table_name} WHERE ReferenceID = ?"
                cursor_py.execute(sql_py, (ref_id,))
            py_rows = cursor_py.fetchall()

            # Query from SSIS (MoneyWaveRemit)
            if table_name == "NegativeListFilter":
                sql_ssis = f"SELECT {', '.join(columns)} FROM MoneyWaveRemit.dbo.{table_name} WHERE ID IN (SELECT ID FROM MoneyWaveRemit.dbo.NegativeList WHERE ReferenceID = ?)"
                cursor_ssis.execute(sql_ssis, (ref_id,))
            else:
                sql_ssis = f"SELECT {', '.join(columns)} FROM MoneyWaveRemit.dbo.{table_name} WHERE ReferenceID = ?"
                cursor_ssis.execute(sql_ssis, (ref_id,))
            ssis_rows = cursor_ssis.fetchall()

            # Sort rows by stable business key to align them perfectly
            py_rows = sorted(py_rows, key=row_sort_key)
            ssis_rows = sorted(ssis_rows, key=row_sort_key)

            max_len = max(len(ssis_rows), len(py_rows))
            for i in range(max_len):
                s_row = ssis_rows[i] if i < len(ssis_rows) else None
                p_row = py_rows[i] if i < len(py_rows) else None

                s_vals = [str(x) if x is not None else "" for x in s_row] if s_row else [""] * len(columns)
                p_vals = [str(x) if x is not None else "" for x in p_row] if p_row else [""] * len(columns)

                match_vals = []
                mismatched_cols = []
                for col_idx, col_name in enumerate(columns):
                    sv = clean_val(s_vals[col_idx])
                    pv = clean_val(p_vals[col_idx])

                    if col_name == "OriginalSource":
                        s_urls = sorted([x.strip() for x in sv.split(";") if x.strip()])
                        p_urls = sorted([x.strip() for x in pv.split(";") if x.strip()])
                        if s_urls == p_urls:
                            sv = pv

                    if sv == pv:
                        match_vals.append("MATCH")
                    else:
                        match_vals.append("MISMATCH")
                        mismatched_cols.append(col_name)

                if mismatched_cols:
                    mismatches_summary.append({
                        "Sheet": sheet_name,
                        "ReferenceID": ref_id,
                        "Mismatches": mismatched_cols
                    })

                rows.append([ref_id, "SSIS (Old)"] + s_vals)
                rows.append([ref_id, "Python (Staging)"] + p_vals)
                rows.append([ref_id, "Match Status"] + match_vals)
                rows.append([""] * (len(columns) + 2))

        headers = ["RefID_Key", "System_Type"] + columns
        return pd.DataFrame(rows, columns=headers)

    print("Building Comparison Sheet Dataframes...")
    # NegativeList comparison sheets
    df_nl_first = build_comparison_rows(first_ids, "NegativeList", cols_negativelist, "NegativeList First 10")
    df_nl_mid = build_comparison_rows(middle_ids, "NegativeList", cols_negativelist, "NegativeList Middle 10")
    df_nl_last = build_comparison_rows(last_ids, "NegativeList", cols_negativelist, "NegativeList Last 10")

    # NegativeListFilter comparison sheets
    df_nlf_first = build_comparison_rows(first_ids, "NegativeListFilter", cols_filter, "Filter First 10")
    df_nlf_mid = build_comparison_rows(middle_ids, "NegativeListFilter", cols_filter, "Filter Middle 10")
    df_nlf_last = build_comparison_rows(last_ids, "NegativeListFilter", cols_filter, "Filter Last 10")

    # NegativeList_Master comparison sheets
    df_nlm_first = build_comparison_rows(first_ids, "NegativeList_Master", cols_master, "Master First 10")
    df_nlm_mid = build_comparison_rows(middle_ids, "NegativeList_Master", cols_master, "Master Middle 10")
    df_nlm_last = build_comparison_rows(last_ids, "NegativeList_Master", cols_master, "Master Last 10")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "LexisNexis_Incremental_Comparison_Report.xlsx")
    print(f"Writing to Excel file: {output_path}...")
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Table Summary", index=False)
        df_nl_first.to_excel(writer, sheet_name="NegativeList First 10", index=False)
        df_nl_mid.to_excel(writer, sheet_name="NegativeList Middle 10", index=False)
        df_nl_last.to_excel(writer, sheet_name="NegativeList Last 10", index=False)
        df_nlf_first.to_excel(writer, sheet_name="Filter First 10", index=False)
        df_nlf_mid.to_excel(writer, sheet_name="Filter Middle 10", index=False)
        df_nlf_last.to_excel(writer, sheet_name="Filter Last 10", index=False)
        df_nlm_first.to_excel(writer, sheet_name="Master First 10", index=False)
        df_nlm_mid.to_excel(writer, sheet_name="Master Middle 10", index=False)
        df_nlm_last.to_excel(writer, sheet_name="Master Last 10", index=False)

    print("Applying Professional Formatting styles...")
    wb = openpyxl.load_workbook(output_path)
    
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_normal = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_ssis = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_python = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_match = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )

    # Format Summary Table
    ws_summary = wb["Table Summary"]
    ws_summary.views.sheetView[0].showGridLines = True
    for col_idx in range(1, 5):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
    for r in range(2, ws_summary.max_row + 1):
        for c in range(1, 5):
            cell = ws_summary.cell(row=r, column=c)
            cell.font = font_normal
            cell.border = border_thin
            if c == 4:
                if cell.value == "YES":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="385723")
                    cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")
                else:
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")
                    cell.fill = PatternFill(start_color="FCE4D6", fill_type="solid")

    # Format Details Sheets
    detail_sheets = [
        "NegativeList First 10", "NegativeList Middle 10", "NegativeList Last 10",
        "Filter First 10", "Filter Middle 10", "Filter Last 10",
        "Master First 10", "Master Middle 10", "Master Last 10"
    ]

    for sname in detail_sheets:
        ws = wb[sname]
        ws.views.sheetView[0].showGridLines = True
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
        for r in range(2, ws.max_row + 1):
            sys_type = ws.cell(row=r, column=2).value
            if not sys_type:
                continue
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = font_normal
                cell.border = border_thin
                if sys_type == "SSIS (Old)":
                    cell.fill = fill_ssis
                elif sys_type == "Python (Staging)":
                    cell.fill = fill_python
                elif sys_type == "Match Status":
                    cell.font = font_bold
                    if c >= 3:
                        if cell.value == "MATCH":
                            cell.font = Font(name="Segoe UI", size=10, bold=True, color="385723")
                            cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")
                        else:
                            cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")
                            cell.fill = PatternFill(start_color="FCE4D6", fill_type="solid")
                    else:
                        cell.fill = fill_match

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(output_path)
    print("\nComparison Report successfully generated!")
    print(f"=== Total Mismatches Detected: {len(mismatches_summary)} ===")

    cursor_py.close()
    conn_py.close()
    cursor_ssis.close()
    conn_ssis.close()

if __name__ == "__main__":
    main()
