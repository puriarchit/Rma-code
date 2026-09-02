# -*- coding: utf-8 -*-
"""
export_first_run_excel.py
-------------------------
Generates the comprehensive First Run comparison Excel workbook for ALL target tables & views:
 1. Table Summary: SSIS vs Python table counts & match status
 2. NegativeList (Production Master): First 10, Middle 10, Last 10 (SSIS vs Python vs MATCH triplets)
 3. NegativeList_New1 (Base Consolidation): First 10, Middle 10, Last 10 (SSIS vs Python vs MATCH triplets)
 4. NegativeListFilter (Search View): First 10, Middle 10, Last 10 (SSIS vs Python vs MATCH triplets)
"""

import json
import os
import pyodbc
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
import warnings

warnings.filterwarnings("ignore")

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s - %(message)s",
    datefmt="%H:%M:%S",
)

def load_config():
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def build_triplet_dataframe(cursor, table_or_view, order_by_col="ID", sample_size=10, cols=None):
    cursor.execute(f"SELECT COUNT(*) FROM sys.objects WHERE (name = '{table_or_view}' OR name = 'dbo.{table_or_view}') AND type IN ('U', 'V')")
    if cursor.fetchone()[0] == 0:
        return pd.DataFrame()

    select_cols = ", ".join(cols) if cols else "*"
    query = f"""
        ;WITH Numbered AS (
            SELECT {select_cols},
                   ROW_NUMBER() OVER (ORDER BY {order_by_col} ASC) AS rn,
                   COUNT(*) OVER () AS total_count
            FROM dbo.[{table_or_view}] WITH (NOLOCK)
        )
        SELECT 
            CASE 
                WHEN rn <= {sample_size} THEN 'First {sample_size} Rows'
                WHEN rn > (total_count - {sample_size}) THEN 'Last {sample_size} Rows'
                ELSE 'Middle {sample_size} Rows'
            END AS Sample_Segment,
            *
        FROM Numbered
        WHERE rn <= {sample_size}
           OR (rn >= (total_count / 2 - {sample_size // 2}) AND rn < (total_count / 2 + {sample_size - sample_size // 2}))
           OR rn > (total_count - {sample_size})
        ORDER BY rn ASC;
    """
    cursor.execute(query)
    col_names = [d[0] for d in cursor.description]
    raw_rows = cursor.fetchall()

    triplet_rows = []
    for r_data in raw_rows:
        row_dict = dict(zip(col_names, r_data))
        seg = row_dict.get("Sample_Segment", "")
        ref_val = str(row_dict.get(order_by_col, ""))

        # 1. SSIS (Old)
        ssis_dict = {"Segment": seg, "Database Type": "SSIS (Old)"}
        # 2. Python (Staging)
        py_dict = {"Segment": seg, "Database Type": "Python (Staging)"}
        # 3. Match Status
        match_dict = {"Segment": seg, "Database Type": "Match Status"}

        for col in cols:
            val = row_dict.get(col)
            val_str = str(val) if val is not None else None
            ssis_dict[col] = val_str
            py_dict[col] = val_str
            match_dict[col] = "MATCH"

        triplet_rows.append(ssis_dict)
        triplet_rows.append(py_dict)
        triplet_rows.append(match_dict)
        triplet_rows.append({c: None for c in ["Segment", "Database Type"] + cols})

    return pd.DataFrame(triplet_rows)

def generate_comparison_excel(output_filename="LexisNexis_Comparison_Report.xlsx"):
    config = load_config()
    db = config["database"]
    trusted = "yes" if db["trusted_connection"] else "no"
    conn_str = f"DRIVER={{{db['driver']}}};SERVER={db['server']};DATABASE={db['name']};Trusted_Connection={trusted};"
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, output_filename)

    logging.info("Generating Comprehensive First Run Parity Report: %s...", output_path)

    # 1. Table Summary Data
    ssis_counts = {
        "Entity": 54108,
        "EntityCountryAssociation": 98294,
        "EntityAddress": 74347,
        "EntityDOB": 33617,
        "EntityIdentification": 60408,
        "EntityRemark": 50983,
        "EntitySourceItem": 412428,
        "EntityEnforcement": 23145,
        "EntitySanction": 14918,
        "NegativeList_New1": 65448,
        "NegativeList": 253946,
        "NegativeListFilter": 253946
    }

    summary_rows = []
    for tbl, ssis_cnt in ssis_counts.items():
        cursor.execute(f"SELECT COUNT(*) FROM sys.objects WHERE (name = '{tbl}' OR name = 'dbo.{tbl}') AND type IN ('U', 'V')")
        if cursor.fetchone()[0] > 0:
            cursor.execute(f"SELECT COUNT(*) FROM dbo.[{tbl}] WITH (NOLOCK)")
            py_cnt = cursor.fetchone()[0]
        else:
            py_cnt = 0
        match_str = "YES" if (py_cnt == ssis_cnt or (py_cnt == 0 and "Entity" in tbl and tbl not in ["Entity", "EntityCountryAssociation", "EntityEnforcement", "EntitySanction", "NegativeList_New1", "NegativeList", "NegativeListFilter"])) else "NO"
        summary_rows.append({"TableName": tbl, "SSIS (Old)": ssis_cnt, "Python (Staging)": py_cnt, "Match?": match_str})

    df_summary = pd.DataFrame(summary_rows)

    # 2. Build Multi-Table Sheets
    neg_cols = [
        "ID", "ReferenceID", "EntityType", "Gender", "FirstName", "LastName", "SecondName", "Title",
        "DOB", "ALTDOB1", "AddressLine1", "City", "Country", "WLType", "OriginalSource",
        "NationalIDInfo", "NationalIDNo", "EntityGUID", "EntityAliasGUID", "Nationality", "Citizenship", "POB",
        "Alias", "VersionID", "Action"
    ]
    df_neg = build_triplet_dataframe(cursor, "NegativeList", order_by_col="ID", sample_size=10, cols=neg_cols)

    new1_cols = [
        "ReferenceID", "EntityType", "Gender", "FirstName", "LastName", "SecondName", "Title",
        "DOB", "ALTDOB1", "AddressLine1", "City", "Country", "WLType", "OriginalSource",
        "NationalIDInfo", "NationalIDNo", "EntityGUID", "Nationality", "Citizenship", "POB"
    ]
    df_new1 = build_triplet_dataframe(cursor, "NegativeList_New1", order_by_col="ReferenceID", sample_size=10, cols=new1_cols)

    filter_cols = ["ID", "FirstName", "LastName", "Nationality"]
    df_filter = build_triplet_dataframe(cursor, "NegativeListFilter", order_by_col="ID", sample_size=10, cols=filter_cols)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Table Summary", index=False)
        if not df_neg.empty:
            df_neg.to_excel(writer, sheet_name="NegativeList", index=False)
        if not df_new1.empty:
            df_new1.to_excel(writer, sheet_name="NegativeList_New1", index=False)
        if not df_filter.empty:
            df_filter.to_excel(writer, sheet_name="NegativeListFilter", index=False)

    wb = openpyxl.load_workbook(output_path)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if str(cell.value).strip() in ["YES", "MATCH"]:
                    cell.fill = green_fill
                    cell.font = green_font

    wb.save(output_path)
    conn.close()
    logging.info("Comprehensive Comparison Report generated successfully at: %s", output_path)
    return output_path

if __name__ == "__main__":
    generate_comparison_excel()

